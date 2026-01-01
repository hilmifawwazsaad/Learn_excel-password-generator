import sys
import random
import string
from openpyxl import load_workbook
from pathlib import Path


def generate_password(length, password_type):
    """
    Generate password berdasarkan tipe dan panjang yang ditentukan
    
    Args:
        length (int): Panjang password
        password_type (str): Tipe password (alfabet/numerik/alfanumerik)
    
    Returns:
        str: Password yang di-generate
    """
    if password_type == "alfabet":
        characters = string.ascii_letters  # a-z, A-Z
    elif password_type == "numerik":
        characters = string.digits  # 0-9
    elif password_type == "alfanumerik":
        characters = string.ascii_letters + string.digits  # a-z, A-Z, 0-9
    else:
        raise ValueError("Tipe password tidak valid! Gunakan: alfabet/numerik/alfanumerik")
    
    # Generate password secara random
    password = ''.join(random.choice(characters) for _ in range(length))
    return password


def process_excel(input_file, password_length, password_type):
    """
    Process Excel file: baca email dari kolom 1 dan generate password di kolom 2
    
    Args:
        input_file (str): Path file Excel input
        password_length (int): Panjang password
        password_type (str): Tipe password
    """
    try:
        # Load workbook
        print(f"Membaca file: {input_file}")
        wb = load_workbook(input_file)
        ws = wb.active
        
        # Hitung jumlah baris yang ada data (skip header jika ada)
        row_count = 0
        start_row = 1
        
        # Check apakah baris pertama adalah header
        first_cell = ws.cell(1, 1).value
        if first_cell and isinstance(first_cell, str) and '@' not in first_cell:
            # Kemungkinan header
            start_row = 2
            print("Header terdeteksi, memulai dari baris 2")
        
        # Generate password untuk setiap baris yang memiliki email
        for row in range(start_row, ws.max_row + 1):
            email = ws.cell(row, 1).value
            
            if email and str(email).strip():  # Jika ada email
                # Generate password
                password = generate_password(password_length, password_type)
                
                # Tulis password di kolom 2
                ws.cell(row, 2, password)
                row_count += 1
                print(f"Row {row}: {email} -> Password generated")
        
        # Simpan ke file baru
        input_path = Path(input_file)
        output_file = input_path.parent / f"result-{input_path.name}"
        wb.save(output_file)
        
        print(f"\nBerhasil! {row_count} password telah di-generate")
        print(f"File disimpan di: {output_file}")
        
    except FileNotFoundError:
        print(f"Error: File '{input_file}' tidak ditemukan!")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)


def main():
    """
    Main function untuk handle command line arguments
    """
    # Validasi jumlah argument
    if len(sys.argv) != 4:
        print("Cara penggunaan yang benar:")
        print("   python main.py [nama_file_excel] [panjang_password] [tipe_password]")
        print("\nContoh:")
        print("   python main.py data.xlsx 12 alfanumerik")
        print("\nTipe password yang tersedia:")
        print("   - alfabet      : Hanya huruf (a-z, A-Z)")
        print("   - numerik      : Hanya angka (0-9)")
        print("   - alfanumerik  : Huruf dan angka (a-z, A-Z, 0-9)")
        sys.exit(1)
    
    # Parse arguments
    input_file = sys.argv[1]
    
    try:
        password_length = int(sys.argv[2])
        if password_length < 1:
            raise ValueError("Panjang password harus lebih dari 0")
    except ValueError as e:
        print(f"Error: Panjang password harus berupa angka positif!")
        sys.exit(1)
    
    password_type = sys.argv[3].lower()
    
    # Validasi tipe password
    valid_types = ["alfabet", "numerik", "alfanumerik"]
    if password_type not in valid_types:
        print(f"Error: Tipe password tidak valid!")
        print(f"Gunakan salah satu: {', '.join(valid_types)}")
        sys.exit(1)
    
    # Process file
    print("=" * 50)
    print("PASSWORD GENERATOR")
    print("=" * 50)
    print(f"File Input    : {input_file}")
    print(f"Panjang Pass  : {password_length} karakter")
    print(f"Tipe Password : {password_type}")
    print("=" * 50)
    print()
    
    process_excel(input_file, password_length, password_type)


if __name__ == "__main__":
    main()
